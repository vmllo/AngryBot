import discord
from discord.ext import commands
from discord.utils import get
from dotenv import load_dotenv
from collections import Counter
import openpyxl
import xlsxwriter
import time
import array
import os
import re
import pandas as pd
import discord
from discord.ext import commands
from discord.ui import View, Button
from sorter.auto_sort_excel import sortmain,averageGS, sortbyGs,auth,sortbyABC,counter, quickfind
from discord import app_commands
cnames = ["Submission","Discord Name","Family Name","Class","Ap","AAP","DP","Position choice", "Kroggy woggy?","Boats?","PVP?","GS"]
comands = ["help - it wont help you","cmdlookup","setGL <stuff you wanna say to g league people>","reload - doesnt do anything yet","names - lists g league guys","getplayer <name>(gear search)","agetplayer(admin) <name> - admin search the bot will bully you if you arent leadership","getsurvey - gear survey","delete - removes pins and deletes commands and bot stuff","button - g league button","status g league status","update <nameofch> updates lookup commands","lookup <name of lookup command>","updategearsheets- kinda says what it does noob",]
intents = discord.Intents.default()
intents.typing = False
intents.presences = False
intents.message_content = True
user_array = []
lookup_array = []
new_message = ""
numberofthumbsup = 0
bot = commands.Bot(command_prefix='!', intents=intents)
role_id = 1249201320884305983
stuff = 0
script_directory = os.path.dirname(os.path.abspath(__file__))
pattern = re.compile(r'http\S+|www\S+')
intents = discord.Intents.default()
intents.members = True
intents.message_content = True  # if you're reading messages
bot = commands.Bot(command_prefix="!", intents=intents)
GUILD_ID = 808051243183767582

@bot.event
async def on_ready():
    with open("usertextfile.txt",'r') as file:
        for line in file.readlines():
            #user_array.append(line.strip())   
            print("") 
    print(f'{bot.user} has connected to Discord!')
    await bot.tree.sync()
    print("Slash commands synced")

emoji1 = bot.get_emoji('👍')
flex = bot.get_emoji('<:Flex:1199972735401607208>')
ping_message = ""
laughingatnoob = discord.utils.get(bot.emojis, name='pointandlaugh')
class YesNo(discord.ui.View):
    def __init__(self, user):
        super().__init__(timeout=30)
        self.user = user
        self.value = None  # True / False

    async def interaction_check(self, interaction):
        return interaction.user == self.user  # only allow original user

    @discord.ui.button(label="Yes", style=discord.ButtonStyle.success)
    async def yes(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.value = True
        await interaction.response.send_message("Okay check Dms")
        self.stop()

    @discord.ui.button(label="No", style=discord.ButtonStyle.danger)
    async def no(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.value = False
        await interaction.response.send_message("Behold in all its glory")
        self.stop()
def clean(name):
    import re, unicodedata
    name = unicodedata.normalize("NFKD", name)
    name = name.encode("ascii","ignore").decode("ascii")
    return re.sub(r'[^a-zA-Z0-9]', '', name).lower()
def cleans(name):
    import re, unicodedata
    name = unicodedata.normalize("NFKD", name)
    name = name.encode("ascii", "ignore").decode("ascii")
    return re.sub(r'[^a-zA-Z0-9 ]', '', name).lower()

@bot.event
async def on_message(message):
    global stuff
    stuff = message
    global new_message
    global numberofthumbsup
    global ping_message
    global lookup_array
    guild = message.guild 
   # guild_get = bot.get_guild(guild.id)
    print(message.content)
    #roles = guild.roles
    #needed_role = get(guild.roles, name="Guild_League_Ping")
    if bot.user.mentioned_in(message):
        await message.channel.send("I am not programmed yet to know what you said, but if its gem you are noob :).\nIf not I'm sure you are a lovely person and if you said something mean go step on a lego.")
    if message.content.startswith('!AB'):
        new_message = message.content.replace('!AB ','')
        new_message =  new_message.lower()
        #print(new_message)
        trigger_message = new_message[0:6]
        display_message = new_message[7:]
        gearlookupmessage = new_message[0:9]
        getlookupmessagen = new_message[10:]
        getgearsurvey = new_message[0:13]
        if trigger_message == "help":
            embed = discord.Embed(title="Commands")
            embed.description = "\n".join(f"{name}" for name in comands)
            await message.channel.send(embed=embed)
        if new_message[0:9] == "cmdlookup":
            xls = pd.read_excel("bdo-resources.xlsx", sheet_name="Sheet1",header=None)
            le = len(xls)
            listforrow = []
            for i in range(0,le):
                rown = xls.iloc[i].dropna().tolist()
                if len(rown) != 2:
                        name = rown[2]
                        name = cleans(name)
                        listforrow.append(name)
            embed = discord.Embed(title="looks up commands Commands\n!AB lookup <one of these commands>")
            embed.description = "\n".join(f"{name}" for name in listforrow)
            await message.channel.send(embed=embed)
        if trigger_message == "setGL":
            messageID = message
            role_id = 1216176901568073848
            #print(messageID)
            await message.channel.send(f"<@&{role_id}>")
            ping_message = await message.channel.send('\n👍: BE READY FOR THE FIRST ROUND\n\n👎: YOU ARENT GOING\n\n🕑 : Gonna be late\n\n💪: You want to chill and be called when we need more\n\n https://tenor.com/view/cat-kitty-kitten-zoom-persuit-gif-17646381782073725526\n ' + display_message)
            await ping_message.pin()
            await ping_message.add_reaction('👍')
            await ping_message.add_reaction('👎')
            await ping_message.add_reaction('🕑')
            await ping_message.add_reaction('💪') 
        if trigger_message == "reload":
            print(ping_message)
        if gearlookupmessage == "getplayer":
            ping_message = await message.channel.send('searching for ' + getlookupmessagen)
            sortmain()
            results = quickfind(getlookupmessagen)
            outgear = []
            if results != "no results":
                if any(role.name == "Officer" or role.name == "GM" or clean(message.author.display_name.lower()) == results[1].lower() or clean(message.author.name.lower()) == results[1].lower()for role in message.author.roles):
                    embed = discord.Embed(title="Player Summary\nFamily name: " + results[2] +"\nDiscord Name: " + results[1] + "\nDo you need to update?: https://forms.gle/fnKMBpFJbPpEYrQq6")
                    lens = len(cnames)
                    for i in range(lens):
                        value = str(cnames[i]) + ": " + str(results[i])
                        outgear.append(value)
                    embed.description = "\n".join(f"{name}" for name in outgear)
                    view = YesNo(message.author)
                    await message.channel.send("Do you want this Dm'd?", view=view) 
                    await view.wait()
                    if view.value == True:
                        await message.author.send(embed=embed)
                    elif view.value == False:
                        await message.channel.send(embed=embed)
                else:
                    embed = discord.Embed(title="Why you care what this doods gear is.. stop being stinky and ask them >:C")
                    await message.channel.send(embed=embed) 
            else:
                embed = discord.Embed(title="No results check spelling please. If you typed noob you are stupid")
                await message.channel.send(embed=embed)
        if new_message[0:10] == "agetplayer":
            results = quickfind(new_message[11:])
            outgear = []
            if results != "no results":
                if any(role.name == "Officer" or role.name == "GM" for role in message.author.roles):
                    embed = discord.Embed(title="Player Summary\nFamily name: " + results[2] +"\nDiscord Name: " + results[1])
                    lens = len(cnames)
                    for i in range(lens):
                        value = str(cnames[i]) + ": " + str(results[i])
                        outgear.append(value)
                    embed.description = "\n".join(f"{name}" for name in outgear)
                    await message.channel.send(embed=embed)
                else:
                    embed = discord.Embed(title="wtf you arent leadership >:C")
                    await message.channel.send(embed=embed) 
            else:
                embed = discord.Embed(title="No results check spelling please. If you typed noob you are stupid")
                await message.channel.send(embed=embed)
        if new_message[0:16] == "updategearsheets":
            sortmain()
            if any(role.name == "Officer" or role.name == "GM" for role in message.author.roles):
                embed = discord.Embed(title="done :3 \nhttps://docs.google.com/spreadsheets/d/1Gggr3CyEMCZcjImj7JGRAEvZZ0TJc3OIpbgTUlvyha4/edit?gid=1765336975#gid=1765336975")
            else: 
                embed = discord.Embed(title="https://forms.gle/fnKMBpFJbPpEYrQq6")
            await message.channel.send(embed=embed)
        if getgearsurvey == "getgearsurvey":
            embed = discord.Embed(title="https://forms.gle/fnKMBpFJbPpEYrQq6")
            await message.channel.send(embed=embed)
        if trigger_message == "getsum":
            embed = discord.Embed(title="Guild Summary")
            auth("Response")
            sortbyGs(1,"NSguildresponses")
            gearave = averageGS(1)
            sortbyABC(3,"Sortedbyclass")
            total = counter(3,14,"Sortedbyclass")
            breakdown = "\n".join(
            f"{cls}: {count}"
            for cls, count in total.items()
            )
            if any(role.name == "Officer" or role.name == "GM" for role in message.author.roles):
                embed.description = (
            f"Average GS:\n {gearave:.2f}\n"
            f"Class breakdown:\n {breakdown}"
            f"\nLink:\n https://docs.google.com/spreadsheets/d/1Gggr3CyEMCZcjImj7JGRAEvZZ0TJc3OIpbgTUlvyha4/edit?gid=1765336975#gid=1765336975")
            else:
                embed.description = (
            f"Average GS:\n {gearave:.2f}\n"
            f"Class breakdown:\n {breakdown}"
            f"\nUpdate gear:\n https://forms.gle/fnKMBpFJbPpEYrQq6"
            )
            await message.channel.send(embed=embed)
        if trigger_message == "button":
            role_id = 1216176901568073848
            class RoleButton(discord.ui.View):
                def __init__(self):
                    super().__init__(timeout=None)
                @discord.ui.button(label="Join Role", style=discord.ButtonStyle.green)
                async def join_role(self, interaction: discord.Interaction, button: discord.ui.Button):
                    role = interaction.guild.get_role(role_id)

                    if role is None:
                        return await interaction.response.send_message(
                            "Role not found!", ephemeral=True
                        )
                    try:
                        await interaction.user.add_roles(role)
                        await interaction.response.send_message(
                        f"You now have the **{role.name}** role!",
                        ephemeral=True
                    )
                    except discord.Forbidden:
                        await interaction.response.send_message(
                            "I don't have permission to assign that role.",
                            ephemeral=True
                        )
             # Send the message with the button
            await message.channel.send(
            "Click to join the role:",
            view=RoleButton()
            )
        if trigger_message == "names":
            role_id = 1198827166088052758
            role = guild.get_role(role_id) 
            print(role)
            with open("C:\\Users\\vwalk\\sorter\\names.txt", "w", encoding="utf-8") as f:
                for member in role.members:
                    f.write(member.name + "\n")
        if trigger_message == "delete":
            ping_message = await message.channel.send('volc protocol starting.. DELETING MESSAGES BEEP BOOP')
            await message.channel.purge(limit=None, check=lambda msg: msg.pinned)
            await ping_message.delete()
            channels = message.channel
            async for msg in channels.history():
                if msg.type is discord.MessageType.pins_add:
                    await msg.delete()
                if(msg.content[0:3]) == "!AB":
                    await msg.delete()
                if msg.author.name == "Angry_Bot":
                    await msg.delete()
        if trigger_message == "status":
            embed = discord.Embed(title="Team Stinkersw")
            temp_array = user_array
            numberofthumbsup = len(temp_array)
            if(numberofthumbsup <= 0):
                with open("usertextfile.txt", "r") as f:
                    temp_array = [line.strip() for line in f if line.strip()]   
            for s in temp_array:
                if s == "Angry_Bot":
                    temp_array.remove("Angry_Bot")
            numberofthumbsup = len(temp_array)
            ping_message = await message.channel.send(str(numberofthumbsup) + "/10")
            embed.description = "\n".join(f"• {name}" for name in temp_array)
            await message.channel.send(embed=embed)
        if trigger_message == "update":
            channels = guild.channels
            workbook = xlsxwriter.Workbook(display_message+'.xlsx')
            worksheet = workbook.add_worksheet()
            display_message = new_message[7:]
            for channel in channels:
                if(channel.name != "NetSlum"):
                    if(channel.name != "BDO Stuff"):
                        if(channel.name != "NSFW"):
                            if(channel.name != "Voice Channels"):
                                if(channel.name != "archive"):
                                    if(channel.name != "war-shit"):
                                        if(channel.name != "Welcome"):
                                            if(channel.name != "General"):
                                                if(channel.name != "class-help"):
                                                    if(channel.name != "Admin"): 
                                                        if(channel.name != "BDO Voice Channels"):
                                                            if(channel.name != "NetSlum BDO"):
                                                                if(channel.name != "other-games"):
                                                                    if(channel.name != "Leadership"):
                                                                        i = 1
                                                                        async for msg in channel.history(): 
                                                                            if channel.name == str(display_message): 
                                                                                stuff = 'A'+ str(i) 
                                                                                stuff2 = 'B' + str(i)
                                                                                stuff3 = 'C' + str(i)
                                                                                worksheet.write(stuff, str(msg.id)) 
                                                                                worksheet.write(stuff2, str(channel.name))
                                                                                worksheet.write(stuff3, str(re.sub(pattern,'',msg.content)))
                                                                                i = i + 1   
            workbook.close()         
            ping_message = await message.channel.send('Done :3')                             
        if trigger_message == "lookup":
            channels = guild.channels
            dm = new_message[7:26] 
            #threads = await channels.threads()
            #ping_message = await message.channel.fetch_message(msg_id)
            #await message.channel.send("stuff" + message)
            files = [f for f in os.listdir(script_directory) if f.endswith('.xlsx')]
            for file in files:
                file_path = os.path.join(script_directory, file)
                df = openpyxl.load_workbook(file_path)
                sh = df.active
                max_rows = sh.max_row
                for move in range(1,max_rows+1):
                    valuez = str(sh.cell(row=move, column=3).value)
                    valuez = valuez.lstrip()
                    valuez = clean(valuez)
                    if str(valuez).lower() == clean(display_message.lower()):
                        print(f"Found row {move} with value {sh.cell(row=move, column=3).value}")   
                        dm = sh.cell(row=move, column=1).value
            for channel in channels:
                #if channel.name == "bdo-resources":
                if(channel.name != "NetSlum"):
                    if(channel.name != "BDO Stuff"):
                        if(channel.name != "NSFW"):
                            if(channel.name != "Voice Channels"):
                                if(channel.name != "archive"):
                                    if(channel.name != "war-shit"):
                                        if(channel.name != "Welcome"):
                                            if(channel.name != "General"):
                                                if(channel.name != "class-help"):
                                                    if(channel.name != "archive"): 
                                                        if(channel.name != "Admin"): 
                                                            if(channel.name != "BDO Voice Channels"):
                                                                if(channel.name != "NetSlum BDO"):
                                                                    if(channel.name != "other-games"):
                                                                        if(channel.name != "Leadership"):  
                                                                            async for msg in channel.history():
                                                                                if str(msg.id) == str(dm):  
                                                                                    await message.channel.send(msg.jump_url)
                                                                                    return
            xls = pd.read_excel("bdo-resources.xlsx", sheet_name="Sheet1",header=None)
            le = len(xls)
            listforrow = []
            for i in range(0,le):
                rown = xls.iloc[i].dropna().tolist()
                if len(rown) != 2:
                        name = rown[2]
                        name = cleans(name)
                        listforrow.append(name)
            embed = discord.Embed(title="looks up commands Commands\n!AB lookup <one of these commands>")
            embed.description = "\n".join(f"{name}" for name in listforrow)
            await message.channel.send(embed=embed)
@bot.event 
async def on_reaction_add(reaction,user):
    global user_array
    global numberofthumbsup
    global stuff
    global flag
    print(user.name)
    if reaction.message.author == bot.user:
        if reaction.emoji == '👍':
            if user.name in user_array: 
                print("")
            else: 
                user_array.append(user.name)
                print(user_array)
                with open('usertextfile.txt', 'r') as f:
                    lines = list(filter(lambda x: x.strip() != '', f.readlines()))
                with open('usertextfile.txt', 'w') as f:
                    f.writelines(lines)
                f.close
                f = open("usertextfile.txt","w+")
                for i in user_array:
                    f.write(i+"\n")
                f.close
            if len(user_array) > 1:
                await ping_message.remove_reaction(reaction.emoji, bot.user)
            numberofthumbsup = len(user_array) - 1
            member = reaction.message.author
            role = get(member.guild.roles, id=role_id)
            await user.add_roles(role)


@bot.tree.command(name="pgetplayer", description="Get player summary",guild=discord.Object(id=GUILD_ID))
@app_commands.describe(name="Discord name to look up")
async def getpla(interaction: discord.Interaction, name: str):

    # ---- LOOKUP ----
    results = quickfind(name)
    outgear = []

    if results == "no results":
        await interaction.response.send_message(
            "No results found. Check spelling.",
            ephemeral=True
        )
        return

    # ---- PERMISSION CHECK ----
    allowed_roles = {"officer", "gm"}
    author_roles = {role.name.lower() for role in interaction.user.roles}

    if (
        interaction.user.name.lower() != results[1].lower()
        and not (allowed_roles & author_roles)
    ):
        await interaction.response.send_message(
            "Why you care what this dude’s gear is… ask them 😤",
            ephemeral=True
        )
        return

    # ---- BUILD EMBED ----
    embed = discord.Embed(
        title=(
            "Player Summary\n"
            f"Family Name: {results[2]}\n"
            f"Discord Name: {results[1]}"
        )
    )

    for i in range(len(cnames)):
        outgear.append(f"{cnames[i]}: {results[i]}")

    embed.description = "\n".join(outgear)

    # ---- PRIVATE RESPONSE ----
    await interaction.response.send_message(
        embed=embed,
        ephemeral=True
    )

@bot.event
async def on_raw_reaction_remove(reaction):
    user = await bot.fetch_user(reaction.user_id)
    global user_array
    global numberofthumbsup
    global flag
    numberofthumbsup = len(user_array) - 1
    # Ignore bot's own reactions
    print(reaction.emoji)
    if str(reaction.emoji) == '👍':
        print("nice thumbs up")
        if numberofthumbsup > 0:
            if str(user) in user_array:
                user_array.remove(str(user))
                print(len(user_array))
                if len(user_array) <= 1:
                    await ping_message.add_reaction(reaction.emoji)
                print(user_array)
                numberofthumbsup = len(user_array) - 1
                f = open("usertextfile.txt","w+")
                for i in user_array:
                    f.write("\n" + i)
                f.close
                with open('usertextfile.txt', 'r') as f:
                    lines = list(filter(lambda x: x.strip() != '', f.readlines()))
                with open('usertextfile.txt', 'w') as f:
                    f.writelines(lines)
                f.close
load_dotenv()
TOKEN = os.getenv('DISCORD_TOKEN')

bot.run(TOKEN)